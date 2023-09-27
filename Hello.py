import openpyxl 
import os 
import tkinter as tk 
from tkinter import messagebox, StringVar 

current_row = 1  # 全局變數，用於追踪目前正在寫入的行

def create_excel(): 
    global current_row  # 使用全局變數
    file_name = "new_blank_workbook.xlsx"

    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)  # 如果文件存在，加載它
        ws = wb.active
    else:
        print("創建一個新的工作簿...") 
        wb = openpyxl.Workbook() 
        ws = wb.active 

    print("向工作簿中添加數據...") 
    ws[f"A{current_row}"].value = selected_option.get() 
    ws[f"B{current_row}"].value = f"=LEFT(A{current_row},1)" 
    current_row += 1  # 在選擇後增加行數

    print("保存工作簿到一個文件...") 
    wb.save(file_name) 
    print(f"保存工作簿到 {os.getcwd()}\\{file_name}...") 

    print("打開新創建的工作簿...") 
    os.startfile(file_name) 


window = tk.Tk()
window.title("軟體視窗")
window.geometry("280x400+{}+{}".format(window.winfo_screenwidth() // 2 - 200, window.winfo_screenheight() // 2 - 200))


# 在視窗中加入一個標籤，提示用戶輸入關鍵字
label = tk.Label(window, text="選擇Support類型", font=("標楷體", 20))
label.place(relx=0.5, rely=0.2, anchor='center')  # 位置設定為視窗的中央
label.pack()  # 設定佈局


options_part1 = ["C015", "C016", "C017", "C019", "C021"]
options_part2 = ["C022", "C023", "C040", "C041", "C046"]
options_part3 = ["C1021", "C1022", "C1121", "C1122", "C1301"]
options_part4 = ["H054", "I080", "I1080", "P085", "P086"]
options_part5 = ["P087", "P093", "P094", "W114", "W115"]
options_part6 = ["W117", "W122", "W124", "W127", "W131"]
options_part7 = ["W132", "W140", "W141"]

options = options_part1 + options_part2 + options_part3 + options_part4 + options_part5 + options_part6 + options_part7


selected_option = StringVar(window)
selected_option.set(options[0])

dropdown = tk.OptionMenu(window, selected_option, *options)
dropdown.place(relx=0.5, rely=0.6, anchor='center')

button = tk.Button(window, text="創建Excel", command=create_excel)
button.place(relx=0.5, rely=0.8, anchor='center')

window.mainloop()
