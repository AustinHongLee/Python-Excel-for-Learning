import openpyxl
import difflib
import sys
import re
import codecs

sys.stdout = codecs.getwriter("utf-8")(open('output.txt', 'w'))
sys.stdout = open('output.txt', 'w')
import re

def exact_match_start(cell_value, keywords):
    """進行開頭匹配"""
    if cell_value is None:
        return None

    for keyword in keywords:
        if cell_value.startswith(keyword):
            print(f"使用了開頭匹配模組，匹配關鍵字為: {keyword}")
            return keyword
    print("exact_match_start沒有匹配值")
    return None    

def fuzzy_match(cell_value, keywords):
    """進行模糊匹配"""
    if cell_value is None:
        return None
    
    words = re.split(r'[\s,;.()\[\]{}=+-]+', cell_value)
    print(f"re.split的結果為: {words}")
    for word in words:
        for keyword in keywords:
            ratio = difflib.SequenceMatcher(None, word.lower(), keyword.lower()).ratio()
            if ratio > 0.8:
                print(f"正在檢查單詞 {word} 與關鍵字 {keyword} 的相似度，相似度為 {ratio}")
            if ratio > 0.9:
                print(f"使用了模糊匹配模組，匹配關鍵字為: {keyword}")
                return keyword
    return None

def in_match(cell_value, keywords):
    """進行整體匹配"""
    if cell_value is None:
        return None

    for keyword in keywords:
        if keyword.lower() in cell_value.lower():
            print(f"使用了整體匹配模組，匹配關鍵字為: {keyword}")
            return keyword
    print("in_match沒有匹配值")
    return None 

# 載入Excel文件
print("正在嘗試加載Excel文件...")
file_path = (
    r'C:\Users\a0976\OneDrive\文件\GitHub\Python Excel for Learning'
    r'\example.xlsx'
)
wb = openpyxl.load_workbook(file_path)
ws = wb.active
print("Excel文件成功加載。")

# 模糊匹配的關鍵字列表
# 分類關鍵字
pipes_and_fittings = [
    "Pipe", "Elbow", "Reducer", "Tee", "Nipple", "Union", "Coupling", "Plug", 
    "Cap", "Stub End", "Swage Nipple", "Concentric Swage", "Eccentric Swage", 
    "Cross", "Olet", "Bend", "Miter Bend", "Lateral"
]

flanges = [
    "Weld Neck Flange", "Socket Weld Flange", "Slip-On Flange", "Lap Joint Flange", 
    "Threaded Flange", "Blind Flange", "Orifice Flange", "Spectacle Blind", 
    "Spacer", "Ring Joint Flange", "LWN Flange", "Square Flange"
]

valves = [
    "Butterfly Valve", "Ball Valve", "Gate Valve", "Dual Plate Check Valve", 
    "Globe Valve", "Needle Valve", "Check Valve", "Diaphragm Valve", "Plug Valve", 
    "Relief Valve", "Safety Valve", "Solenoid Valve", "Pressure Reducing Valve"
]

control_valves = [
    "Globe Control Valve", "Butterfly Control Valve", "Ball Control Valve", 
    "Diaphragm Control Valve", "Pinch Control Valve", "Gate Control Valve", 
    "Rotary Control Valve", "Self-operated Control Valve", "Pilot-operated Control Valve",
    "Sliding Cylinder Control Valve", "Flapper-nozzle Control Valve", 
    "Jet Pipe Control Valve", "Eccentric Disk Control Valve", "Angle Control Valve"
]

bolts_and_gaskets = [
    "Bolt", "Nut", "Gasket", "Ring Gasket", "Spiral Wound Gasket", 
    "Metallic Gasket", "Non-Metallic Gasket"
]

misc = [
    "Support", "Sockolet", "Weldolet", "Expansion Joint", "Strainer", 
    "Blind", "Spacer", "Rupture Disc"
]

# 組合所有關鍵字
keywords = pipes_and_fittings + flanges + valves + control_valves + bolts_and_gaskets + misc


print("開始迭代工作表中的{ws.max_row}行...")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
    cell = row[0]
    try:
        print(f"正在檢查第{cell.row}行的單元格，其值為: {cell.value}")
    except UnicodeEncodeError:
        print(f"正在檢查第{cell.row}行的單元格，但無法輸出其值")
    #...[其他代碼不變]


    matched_keyword = exact_match_start(cell.value, keywords) or fuzzy_match(cell.value, keywords) or in_match(cell.value, keywords)

    if matched_keyword:
        print(f"在第{cell.row}行的單元格中找到了匹配項 {matched_keyword}。正在將值寫入B列{cell.row}行。")
        ws.cell(row=cell.row, column=2).value = matched_keyword
    else:
        ws.cell(row=cell.row, column=2).value = "-"
        print(f"在第{cell.row}行的單元格中沒有找到匹配項。")

# 保存Excel文件
print("正在保存Excel文件...")
wb.save(file_path)
print("Excel文件保存成功。")

# 讀取輸出文件並輸出到控制台
with open('output.txt', 'r') as f:
    print(f.read())